import openpyxl
import random
import string
import nltk
nltk.download('words')
from nltk.corpus import words

class ColumnarCipher:
    
    def __init__(self, key):
        
        self.key = key

    def encrypt(self, plaintext):
        
        plaintext = plaintext.upper().replace(" ", "")
        num_columns = len(self.key)
        num_rows = (len(plaintext) + num_columns - 1) // num_columns
        padded_plaintext = plaintext + " " * (num_rows * num_columns - len(plaintext))
        matrix = []
        for i in range(num_rows):
            row = [padded_plaintext[j] for j in range(i * num_columns, (i + 1) * num_columns)]
            matrix.append(row)
        sorted_key = sorted(range(num_columns), key=lambda x: self.key[x])
        rearranged_matrix = [[matrix[row][col] for col in sorted_key] for row in range(num_rows)]
        ciphertext = ""
        for i in range(num_columns):
            for j in range(num_rows):
                ciphertext += rearranged_matrix[j][i]
        return ciphertext

    def decrypt(self, ciphertext):
        
        num_columns = len(self.key)
        num_rows = (len(ciphertext) + num_columns - 1) // num_columns
        matrix = [[""] * num_columns for _ in range(num_rows)]
        sorted_key = sorted(range(num_columns), key=lambda x: self.key[x])
        index = 0
        for col in sorted_key:
            for row in range(num_rows):
                matrix[row][col] = ciphertext[index]
                index += 1
        plaintext = ""
        for row in range(num_rows):
            for col in range(num_columns):
                plaintext += matrix[row][col]
        return plaintext

def generate_plaintext():
    
    word = random.choice(words.words())
    return ''.join(random.choices(word.upper(), k=len(word)))

key = "CRYPTO"
columnar = ColumnarCipher(key)

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "Plain Text"
sheet["B1"] = "Key"
sheet["C1"] = "Cipher Text"
sheet["D1"] = "Decrypted Text"

for i in range(2, 10002):
    plaintext = generate_plaintext()
    encrypted_message = columnar.encrypt(plaintext)
    decrypted_message = columnar.decrypt(encrypted_message)
    sheet.cell(row=i, column=1, value=plaintext)
    sheet.cell(row=i, column=2, value=key)
    sheet.cell(row=i, column=3, value=encrypted_message)
    sheet.cell(row=i, column=4, value=decrypted_message)

workbook.save("COLUMNAR_TRANSPOSITION_CIPHER_DATASET.xlsx")
