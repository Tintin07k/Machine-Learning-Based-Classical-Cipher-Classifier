import openpyxl
import random
import string
import nltk
nltk.download('words')
from nltk.corpus import words

class ScytaleCipher:
    
    def __init__(self, diameter):
        
        self.diameter = diameter

    def encrypt(self, plaintext):
        
        num_rows = (len(plaintext) + self.diameter - 1) // self.diameter
        padded_plaintext = plaintext.ljust(num_rows * self.diameter)
        encrypted_message = ""
        for col in range(self.diameter):
            for row in range(num_rows):
                encrypted_message += padded_plaintext[row * self.diameter + col]

        return encrypted_message

    def decrypt(self, ciphertext):
        
        num_rows = (len(ciphertext) + self.diameter - 1) // self.diameter
        decrypted_message = ""
        for row in range(num_rows):
            for col in range(self.diameter):
                index = col * num_rows + row
                if index < len(ciphertext):
                    decrypted_message += ciphertext[index]

        return decrypted_message

def generate_plaintext():
    
    word = random.choice(words.words())
    return ''.join(random.choices(word.upper(), k=len(word)))

diameter = 5
scytale = ScytaleCipher(diameter)

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1, value="Plain Text")
sheet.cell(row=1, column=2, value="Key (Diameter)")
sheet.cell(row=1, column=3, value="Cipher Text")
sheet.cell(row=1, column=4, value="Decrypted Text")

for i in range(2, 12):
    plaintext = generate_plaintext()
    encrypted_message = scytale.encrypt(plaintext)
    decrypted_message = scytale.decrypt(encrypted_message)
    sheet.cell(row=i, column=1, value=plaintext)
    sheet.cell(row=i, column=2, value=diameter)
    sheet.cell(row=i, column=3, value=encrypted_message)
    sheet.cell(row=i, column=4, value=decrypted_message)

workbook.save("SCYTALE_CIPHER_DATASET.xlsx")
