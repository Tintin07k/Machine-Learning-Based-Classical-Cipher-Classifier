import openpyxl
import random
import string
import nltk
nltk.download('words')
from nltk.corpus import words

class GronsfeldCipher:
    
    def __init__(self, key):
        
        self.key = key

    def encrypt(self, plaintext):
        
        plaintext = plaintext.upper()
        encrypted_message = ''
        for i in range(len(plaintext)):
            char = plaintext[i]
            shift = int(self.key[i % len(self.key)]) 
            if char.isalpha():
                encrypted_char = chr((ord(char) - ord('A') + shift) % 26 + ord('A'))
            else:
                encrypted_char = char
            encrypted_message += encrypted_char
        return encrypted_message

    def decrypt(self, ciphertext):
        
        ciphertext = ciphertext.upper()
        decrypted_message = ''
        for i in range(len(ciphertext)):
            char = ciphertext[i]
            shift = int(self.key[i % len(self.key)])  
            if char.isalpha():
                decrypted_char = chr((ord(char) - ord('A') - shift) % 26 + ord('A'))
            else:
                decrypted_char = char
            decrypted_message += decrypted_char
        return decrypted_message

def generate_plaintext():
    
    word = random.choice(words.words())
    return ''.join(random.choices(word.upper(), k=len(word)))

key = '31415'  
gronsfeld = GronsfeldCipher(key)

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet["A1"] = "Plain Text"
sheet["B1"] = "Key" 
sheet["C1"] = "Cipher Text"
sheet["D1"] = "Decrypted Text"

for i in range(2, 10002):
    plaintext = generate_plaintext()
    encrypted_message = gronsfeld.encrypt(plaintext)
    decrypted_message = gronsfeld.decrypt(encrypted_message)
    sheet.cell(row=i, column=1, value=plaintext)
    sheet.cell(row=i, column=2, value=key)
    sheet.cell(row=i, column=3, value=encrypted_message)
    sheet.cell(row=i, column=4, value=decrypted_message)

workbook.save("GRONSFELD_CIPHER_DATASET.xlsx")
