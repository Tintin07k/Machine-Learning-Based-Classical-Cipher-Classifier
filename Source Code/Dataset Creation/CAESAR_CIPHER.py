import openpyxl
import nltk
import random
from nltk.corpus import words
from openpyxl import Workbook
nltk.download('words')
word_list = words.words()

random_words = random.sample(word_list, k=10000)

def encrypt(plaintext, shift):
    
    ciphertext = ""
    for char in plaintext:
        if char.isalpha():
            ascii_offset = 65 if char.isupper() else 97
            encrypted_char = chr((ord(char) - ascii_offset + shift) % 26 + ascii_offset)
            ciphertext += encrypted_char
        else:
            ciphertext += char
    return ciphertext


def decrypt(ciphertext, shift):
    
    plaintext = ""
    for char in ciphertext:
        if char.isalpha():
            ascii_offset = 65 if char.isupper() else 97
            decrypted_char = chr((ord(char) - ascii_offset - shift) % 26 + ascii_offset)
            plaintext += decrypted_char
        else:
            plaintext += char
    return plaintext

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "Plain Text"
sheet["B1"] = "Shift"
sheet["C1"] = "Ciphert Text"
sheet["D1"] = "Decrypted Text"

row = 2
count = 0
for word in random_words:
    for shift in range(-25, 26):
        plaintext = word
        ciphertext = encrypt(plaintext, shift)
        decrypted_plaintext = decrypt(ciphertext, shift)
        sheet.cell(row=row, column=1, value=plaintext)
        sheet.cell(row=row, column=2, value=shift)
        sheet.cell(row=row, column=3, value=ciphertext)
        sheet.cell(row=row, column=4, value=decrypted_plaintext)
        row += 1
    count += 1
    if count == 200:
        break

workbook.save("CAESAR_CIPHER_DATASET.xlsx")
